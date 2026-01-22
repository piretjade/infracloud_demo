#!/usr/bin/env python3
"""
Cisco CSR1000v / IOS XE 16.9+ NETCONF Inventory Script
Collects device characteristics, hardware inventory, interfaces,
routing (OSPF), and neighbor info (CDP/LLDP) into an Excel file.
"""

from ncclient import manager
from xml.dom.minidom import parseString
from openpyxl import Workbook

# --- Router details ---
ROUTER = {
    "host": "devnetsandboxiosxec9k.cisco.com",
    "port": 830,
    "user": "jade.piret",
    "password": "43JxCmOSInJU2_w_",
}

# --- Connect to router ---
print("\nConnecting to router via NETCONF...\n")
m = manager.connect(
    host=ROUTER["host"],
    port=ROUTER["port"],
    username=ROUTER["user"],
    password=ROUTER["password"],
    hostkey_verify=False,
    device_params={"name": "csr"},
    look_for_keys=False,
    allow_agent=False
)

# --- Helper functions ---
def netconf_get(filter_xml):
    reply = m.get(filter=("subtree", filter_xml))
    return parseString(reply.data_xml).toprettyxml()

def netconf_get_config(filter_xml):
    reply = m.get_config(source="running", filter=("subtree", filter_xml))
    return parseString(reply.data_xml).toprettyxml()

# === Define filters ===
filters = {
    "Hostname": """
        <native xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-native">
            <hostname/>
        </native>
    """,
    "Platform_Software_Version": """
        <platform-software-version xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-platform-software-oper"/>
    """,
    "Hardware_Inventory": """
        <components xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-platform-oper"/>
    """,
    "Interfaces": """
        <interfaces xmlns="urn:ietf:params:xml:ns:yang:ietf-interfaces">
            <interface/>
        </interfaces>
    """,
    "Routing_OSPF": """
        <ospf-oper-data xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-ospf-oper"/>
    """,
    "CDP_Neighbors": """
        <cdp-neighbor-details xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-cdp-oper"/>
    """,
    "LLDP_Neighbors": """
        <lldp-entries xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-lldp-oper"/>
    """
}

# === Create Excel workbook ===
wb = Workbook()
ws = wb.active
ws.title = "Hostname"

# === Collect and store data ===
for idx, (sheet_name, xml_filter) in enumerate(filters.items()):
    print(f"Collecting: {sheet_name} ...")
    if sheet_name == "Hostname":
        data = netconf_get_config(xml_filter)
    else:
        data = netconf_get(xml_filter)
    
    # Add a worksheet
    if idx != 0:
        ws = wb.create_sheet(title=sheet_name)
    # Write XML output to the sheet
    for i, line in enumerate(data.splitlines(), start=1):
        ws.cell(row=i, column=1, value=line)

# === Save Excel file ===
output_file = "netconf_inventory.xlsx"
wb.save(output_file)

print(f"\n✅ NETCONF inventory complete. Saved to {output_file}\n")

m.close_session()
