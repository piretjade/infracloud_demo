from netmiko import ConnectHandler
from openpyxl import Workbook, load_workbook
import datetime
import os

# Gegevens van het toestel
device = {
    "device_type": "cisco_ios",
    "host": "devnetsandboxiosxec9k.cisco.com",
    "port": 22,
    "username": "jade.piret",
    "password": "43JxCmOSInJU2_w_"
}

print("Current date and time:")
print(datetime.datetime.now())
print("Connecting via SSH => show version")

# SSH-verbinding maken
sshCli = ConnectHandler(**device)

# show version uitvoeren
output = sshCli.send_command("show version")

# Variabelen initialiseren
ios_version = ""
hostname = ""
sys_uptime = ""
num_interfaces = ""

# Output parsen
for line in output.splitlines():
    if "Cisco IOS Software" in line:
        ios_version = line.strip()
    elif "uptime" in line:
        hostname = line.split()[0]
        sys_uptime = line.strip()
    elif "interface" in line and "Virtual" not in line:
        # bv. "2 Virtual Ethernet interfaces"
        num_interfaces = line.split()[0]

print("IOS Version")
print(ios_version)
print("Hostname")
print(hostname)
print("System uptime")
print(sys_uptime)
print("Number of Interfaces")
print(num_interfaces)

# Excel-bestandsnaam
excel_file = "device_info.xlsx"

# Bestaat het Excel-bestand al?
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "DeviceInfo"
    # Kolomtitels
    ws.append(["Datum/tijd", "IOS Version", "Hostname", "System Uptime", "Number of Interfaces"])

# Nieuwe rij toevoegen met data
ws.append([
    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ios_version,
    hostname,
    sys_uptime,
    num_interfaces
])

# Opslaan
wb.save(excel_file)
print(f"Gegevens opgeslagen in {excel_file}")

commands = {
    "show ip protocols": "Routing protocols",
    "show ip interface brief": "IP interfaces (brief)",
    "show ip route": "Routing table",
    "show cdp neighbors": "CDP neighbors"
}

for cmd, beschrijving in commands.items():
    print("\n========================================")
    print(f"{beschrijving} -> {cmd}")
    print("========================================")
    output_cmd = sshCli.send_command(cmd)
    print(output_cmd)

# SSH-verbinding sluiten
sshCli.disconnect()
